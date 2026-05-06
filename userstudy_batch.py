"""
User Study — Batch Recommendation Generation
==============================================
Reads directly from the Google Forms Excel export
(Stage_1__Participant_Input_Form__Responses_.xlsx)
and generates two recommendation playlists for every participant.

Usage
-----
  1. Place the Excel file in the same folder as this script
  2. Run: python userstudy_batch.py
     → generates userstudy_output/<PART-XXX>_recommendations.txt for every participant
  3. Optionally run a single participant:
     python userstudy_batch.py --participant PART-001

Requirements
------------
  pip install requests numpy scikit-learn openpyxl
  All phase2_output/ and phase3_output/ files must be present.
"""

import os
import sys
import pickle
import random
import time
import argparse
import requests
import numpy as np
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"], check=True)
    import openpyxl

try:
    from sklearn.metrics.pairwise import cosine_similarity
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "scikit-learn", "-q"], check=True)
    from sklearn.metrics.pairwise import cosine_similarity

# ────────────────────────────────────────────────
# Configuration
# ────────────────────────────────────────────────
EXCEL_FILE       = os.path.join(os.path.dirname(__file__),
                                "Stage_1__Participant_Input_Form__Responses_.xlsx")
API_KEY          = os.getenv("LASTFM_API_KEY", "f4f4b05dcb5e7186c9762534a06a4bda")
BASE_URL         = "https://ws.audioscrobbler.com/2.0/"
DELAY            = 0.25

PHASE2_DIR       = os.getenv("PHASE2_OUTPUT_DIR",
                              os.path.join(os.path.dirname(__file__), "phase2_output"))
PHASE3_DIR       = os.getenv("PHASE3_OUTPUT_DIR",
                              os.path.join(os.path.dirname(__file__), "phase3_output"))
OUTPUT_DIR       = os.path.join(os.path.dirname(__file__), "userstudy_output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

CONTEXT_WINDOW   = 50
RECENCY_DECAY    = 1.0
TOP_K            = 10
MIN_MATCHED      = 15     # Section 4.6.5.2


# ────────────────────────────────────────────────
# Excel column layout (from Google Forms export)
# ────────────────────────────────────────────────
# The form has 20 song pairs: "Song N - Artist Name" and "Song N - Song Title"
# Song 1 is most recent (as per form instructions)
COL_PARTICIPANT_CODE = "Unique Participant Code"
COL_MOOD             = "When you listen to these songs, what is your general mood?"
SONG_COUNT           = 20


# ────────────────────────────────────────────────
# Load Excel and parse all participants
# ────────────────────────────────────────────────

def load_participants(excel_path: str) -> list[dict]:
    """
    Parse the Google Forms Excel export.
    Returns list of participant dicts:
      {
        'id':     'PART-001',
        'mood':   'Sad / Melancholic',
        'tracks': [('Artist', 'Title'), ...]   ← Song 1 first (most recent)
      }
    """
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}")
        print("Place the Google Forms export in the same folder as this script.")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    # Read header row
    headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Build column index lookup
    col_idx = {h: i for i, h in enumerate(headers)}

    def get(row_vals, col_name):
        idx = col_idx.get(col_name)
        if idx is None:
            return ""
        val = row_vals[idx]
        return str(val).strip() if val is not None else ""

    participants = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        participant_id = get(row, COL_PARTICIPANT_CODE)
        if not participant_id or participant_id == "None":
            continue

        mood = get(row, COL_MOOD)

        # Extract song pairs — Song 1 (most recent) through Song 20
        tracks = []
        for n in range(1, SONG_COUNT + 1):
            artist_col = f"Song {n} - Artist Name"
            title_col  = f"Song {n} - Song Title"
            artist = get(row, artist_col)
            title  = get(row, title_col)
            if artist and title and artist != "None" and title != "None":
                tracks.append((artist, title))

        if tracks:
            participants.append({
                "id":     participant_id,
                "mood":   mood,
                "tracks": tracks,
            })

    wb.close()
    return participants


# ────────────────────────────────────────────────
# Load model outputs (once, shared across all participants)
# ────────────────────────────────────────────────

def load_pickle(path: str):
    if not os.path.exists(path):
        print(f"ERROR: Required file not found: {path}")
        sys.exit(1)
    with open(path, "rb") as f:
        return pickle.load(f)


print("Loading model outputs...")
semantic_emb      = load_pickle(os.path.join(PHASE2_DIR, "semantic_embeddings.pkl"))
emotion_aware_emb = load_pickle(os.path.join(PHASE2_DIR, "emotion_aware_embeddings.pkl"))
track_rep_sem     = load_pickle(os.path.join(PHASE3_DIR, "track_representations_semantic.pkl"))
track_rep_emo     = load_pickle(os.path.join(PHASE3_DIR, "track_representations_emotion.pkl"))
train_events      = load_pickle(os.path.join(PHASE3_DIR, "train_events.pkl"))
print(f"  Semantic embeddings   : {len(semantic_emb):,} tags")
print(f"  Emotion embeddings    : {len(emotion_aware_emb):,} tags")
print(f"  Track reps (semantic) : {len(track_rep_sem):,} tracks")
print(f"  Track reps (emotion)  : {len(track_rep_emo):,} tracks")
print(f"  Train events          : {len(train_events):,}")

# Build track → tags lookup once
print("\nBuilding track-tag lookup...")
track_tag_lookup = defaultdict(set)
for ev in train_events:
    key = f"{ev['artist'].lower()}||{ev['track'].lower()}"
    for tag in ev["tags"]:
        track_tag_lookup[key].add(tag)
track_tag_lookup = dict(track_tag_lookup)
print(f"  Unique tracks in dataset: {len(track_tag_lookup):,}")

# Pre-build candidate matrices for fast cosine similarity (built once)
print("\nPre-building candidate matrices...")
_candidate_keys_sem   = list(track_rep_sem.keys())
_candidate_matrix_sem = np.array([track_rep_sem[k] for k in _candidate_keys_sem], dtype=np.float32)
_candidate_keys_emo   = list(track_rep_emo.keys())
_candidate_matrix_emo = np.array([track_rep_emo[k] for k in _candidate_keys_emo], dtype=np.float32)
print(f"  Semantic matrix  : {_candidate_matrix_sem.shape}")
print(f"  Emotion matrix   : {_candidate_matrix_emo.shape}")


# ────────────────────────────────────────────────
# Last.fm API fallback
# ────────────────────────────────────────────────

_lastfm_cache = {}   # cache API calls across participants to save time

def call_api(params: dict, retries: int = 3) -> dict | None:
    params = dict(params)
    params.update({"api_key": API_KEY, "format": "json"})
    for attempt in range(retries):
        try:
            r = requests.get(BASE_URL, params=params, timeout=10)
            r.raise_for_status()
            data = r.json()
            if "error" in data:
                return None
            return data
        except requests.RequestException:
            time.sleep(2 ** attempt)
    return None


def fetch_tags_from_lastfm(artist: str, track: str) -> list[str]:
    """Fetch tags with in-memory cache to avoid duplicate API calls."""
    cache_key = f"{artist.lower()}||{track.lower()}"
    if cache_key in _lastfm_cache:
        return _lastfm_cache[cache_key]

    time.sleep(DELAY)
    data = call_api({"method": "track.getTopTags", "artist": artist, "track": track})
    if data:
        tags = data.get("toptags", {}).get("tag", [])
        if isinstance(tags, list) and tags:
            result = [t["name"].lower().strip() for t in tags if isinstance(t, dict)]
            _lastfm_cache[cache_key] = result
            return result

    # Artist fallback
    time.sleep(DELAY)
    data = call_api({"method": "artist.getTopTags", "artist": artist})
    if data:
        tags = data.get("toptags", {}).get("tag", [])
        if isinstance(tags, list) and tags:
            result = [t["name"].lower().strip() for t in tags if isinstance(t, dict)]
            _lastfm_cache[cache_key] = result
            return result

    _lastfm_cache[cache_key] = []
    return []


# ────────────────────────────────────────────────
# Track matching
# ────────────────────────────────────────────────

def normalize(s: str) -> str:
    import re
    s = s.lower().strip()
    s = re.sub(r"[^\w\s]", "", s)
    s = re.sub(r"\s+", " ", s)
    return s


def match_track(artist: str, track: str) -> tuple:
    """
    Returns (track_key | None, tags: set, source: str)
    source: 'dataset' | 'lastfm' | 'unmatched'
    """
    key = f"{artist.lower()}||{track.lower()}"
    if key in track_tag_lookup:
        return key, track_tag_lookup[key], "dataset"

    # Normalized fuzzy match
    na, nt = normalize(artist), normalize(track)
    for existing_key in track_tag_lookup:
        ea, et = existing_key.split("||", 1)
        if normalize(ea) == na and normalize(et) == nt:
            return existing_key, track_tag_lookup[existing_key], "dataset"

    # Last.fm fallback
    api_tags   = fetch_tags_from_lastfm(artist, track)
    valid_tags = {t for t in api_tags if t in semantic_emb}
    if valid_tags:
        return key, valid_tags, "lastfm"

    return None, set(), "unmatched"


# ────────────────────────────────────────────────
# Context vector
# ────────────────────────────────────────────────

def build_context_vector(context_tracks: list, embeddings: dict) -> np.ndarray | None:
    """Recency-weighted context vector. Most recent = highest weight."""
    n             = len(context_tracks)
    sorted_tracks = sorted(context_tracks, key=lambda x: x["rank"], reverse=True)
    positions     = np.linspace(-RECENCY_DECAY, 0, n)
    weights       = np.exp(positions)

    weighted_vecs = []
    event_weights = []

    for i, t in enumerate(sorted_tracks):
        tag_vecs = [embeddings[tag] for tag in t["tags"] if tag in embeddings]
        if not tag_vecs:
            continue
        weighted_vecs.append(np.mean(tag_vecs, axis=0))
        event_weights.append(weights[i])

    if not weighted_vecs:
        return None

    ew    = np.array(event_weights)
    stack = np.stack(weighted_vecs)
    return (np.sum(stack * ew[:, np.newaxis], axis=0) / ew.sum()).astype(np.float32)


# ────────────────────────────────────────────────
# Recommendation generation
# ────────────────────────────────────────────────

MAX_TRACKS_PER_ARTIST = 2   # diversity constraint — Section 4.5.3


def generate_recommendations(
    ctx_vec: np.ndarray,
    candidate_keys: list,
    candidate_matrix: np.ndarray,
    exclude_keys: set,
    k: int,
) -> list[tuple]:
    """
    Returns list of (artist, track, score), ranked by cosine similarity.

    A diversity constraint (re-ranking) is applied: no more than
    MAX_TRACKS_PER_ARTIST tracks from the same artist appear in the
    top-k results. This prevents artist-level clustering from dominating
    the playlist when many tracks from one artist share near-identical
    tag representations (Section 4.5.3).
    """
    sims           = cosine_similarity(ctx_vec.reshape(1, -1), candidate_matrix)[0]
    ranked_indices = np.argsort(sims)[::-1]

    results       = []
    artist_counts = defaultdict(int)   # track how many songs per artist

    for idx in ranked_indices:
        ckey = candidate_keys[idx]
        if ckey in exclude_keys:
            continue
        artist, track = ckey.split("||", 1)
        artist_norm   = artist.lower().strip()

        # Diversity filter: skip if artist already has MAX_TRACKS_PER_ARTIST
        if artist_counts[artist_norm] >= MAX_TRACKS_PER_ARTIST:
            continue

        artist_counts[artist_norm] += 1
        results.append((artist.title(), track.title(), float(sims[idx])))

        if len(results) >= k:
            break

    return results


# ────────────────────────────────────────────────
# Process one participant
# ────────────────────────────────────────────────

def process_participant(participant: dict) -> str | None:
    """
    Full pipeline for one participant.
    Returns output text string, or None if participant fails minimum match threshold.
    """
    pid    = participant["id"]
    mood   = participant["mood"]
    tracks = participant["tracks"]

    print(f"\n{'='*55}")
    print(f"Processing {pid}  (mood: {mood})")
    print(f"{'='*55}")

    # ── Match tracks ──────────────────────────────
    matched   = []
    unmatched = []

    for i, (artist, title) in enumerate(tracks):
        key, tags, source = match_track(artist, title)

        if key is not None and tags:
            valid_tags = {t for t in tags if t in semantic_emb}
            if valid_tags:
                matched.append({
                    "rank":   i,   # 0 = most recent
                    "artist": artist,
                    "track":  title,
                    "key":    key,
                    "tags":   valid_tags,
                    "source": source,
                })
                status = f"✓ [{source}]"
            else:
                unmatched.append((artist, title))
                status = "✗ no embeddings"
        else:
            unmatched.append((artist, title))
            status = "✗ unmatched"

        print(f"  {i+1:>2}. {artist[:30]:<30} — {title[:35]:<35} {status}")

    print(f"\n  Matched: {len(matched)} / {len(tracks)}")

    if len(matched) < MIN_MATCHED:
        print(f"  SKIPPED: fewer than {MIN_MATCHED} matched tracks.")
        print(f"  Unmatched: {[f'{a} - {t}' for a,t in unmatched]}")
        return None

    # Use up to CONTEXT_WINDOW most recent matched tracks
    context_tracks = sorted(matched, key=lambda x: x["rank"])[:CONTEXT_WINDOW]

    # ── Build context vectors ─────────────────────
    ctx_sem = build_context_vector(context_tracks, semantic_emb)
    ctx_emo = build_context_vector(context_tracks, emotion_aware_emb)

    if ctx_sem is None or ctx_emo is None:
        print("  SKIPPED: could not build context vectors.")
        return None

    # ── Generate recommendations ──────────────────
    exclude_keys = {m["key"] for m in matched}
    recs_sem = generate_recommendations(ctx_sem, _candidate_keys_sem,
                                         _candidate_matrix_sem, exclude_keys, TOP_K)
    recs_emo = generate_recommendations(ctx_emo, _candidate_keys_emo,
                                         _candidate_matrix_emo, exclude_keys, TOP_K)

    # ── Randomize presentation order ──────────────
    random.seed(hash(pid))
    order = random.sample(["A", "B"], 2)
    playlists = {
        "A": ("Semantic",      recs_sem),
        "B": ("Emotion-Aware", recs_emo),
    }

    # ── Format output ─────────────────────────────
    lines = []
    lines.append("=" * 60)
    lines.append("USER STUDY — RECOMMENDATIONS")
    lines.append(f"Participant    : {pid}")
    lines.append(f"Reported mood  : {mood}")
    lines.append(f"Matched tracks : {len(matched)} / {len(tracks)}")
    lines.append(f"Context tracks : {len(context_tracks)}")
    lines.append("=" * 60)
    lines.append("")
    lines.append("PARTICIPANT'S INPUT TRACKS (most recent first):")
    for m in sorted(context_tracks, key=lambda x: x["rank"]):
        lines.append(f"  {m['rank']+1:>2}. {m['artist']} — {m['track']}  [{m['source']}]")
    lines.append("")

    for label in order:
        model_name, recs = playlists[label]
        lines.append(f"PLAYLIST {label}")
        lines.append("-" * 40)
        for i, (artist, track, score) in enumerate(recs, 1):
            lines.append(f"  {i:>2}. {artist} — {track}")
        lines.append("")

    lines.append("─" * 60)
    lines.append("INTERNAL RECORD (researcher only — do not share):")
    lines.append(f"  Playlist A = {playlists['A'][0]} model")
    lines.append(f"  Playlist B = {playlists['B'][0]} model")
    lines.append(f"  Order shown to participant: {' → '.join(order)}")
    lines.append("─" * 60)

    return "\n".join(lines)


# ────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate recommendations for user study participants")
    parser.add_argument("--participant", type=str, default=None,
                        help="Run for a single participant code e.g. PART-001")
    args = parser.parse_args()

    # Load all participants from Excel
    print(f"\nLoading participants from {EXCEL_FILE}...")
    all_participants = load_participants(EXCEL_FILE)
    print(f"  Found {len(all_participants)} participants")

    # Filter to single participant if specified
    if args.participant:
        target = args.participant.strip().upper()
        all_participants = [p for p in all_participants if p["id"].upper() == target]
        if not all_participants:
            print(f"ERROR: Participant '{args.participant}' not found in Excel file.")
            print(f"Available codes: {[p['id'] for p in load_participants(EXCEL_FILE)]}")
            sys.exit(1)

    # Process each participant
    success_count = 0
    skipped       = []

    for participant in all_participants:
        output_text = process_participant(participant)

        if output_text is None:
            skipped.append(participant["id"])
            continue

        # Save to file
        out_path = os.path.join(OUTPUT_DIR, f"{participant['id']}_recommendations.txt")
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(output_text)

        print(f"\n  ✓ Saved: {out_path}")
        success_count += 1

    # Final summary
    print("\n" + "=" * 55)
    print("BATCH COMPLETE")
    print("=" * 55)
    print(f"  Generated : {success_count} / {len(all_participants)}")
    print(f"  Skipped   : {len(skipped)}")
    if skipped:
        print(f"  Skipped IDs: {skipped}")
    print(f"\n  Output folder: {OUTPUT_DIR}")
    print("  Each file contains two playlists (A and B) for the participant.")
    print("  Share only PLAYLIST A and PLAYLIST B with the participant.")
    print("  Keep the INTERNAL RECORD section for your own reference.")